from __future__ import annotations

from io import BytesIO
from typing import List, Tuple

from fastapi import UploadFile, HTTPException, status

# PDF
from pypdf import PdfReader

# DOCX
from docx import Document

# XLSX
from openpyxl import load_workbook


SUPPORTED_EXT = {"txt", "md", "pdf", "docx", "xlsx", "csv"}


def _get_ext(filename: str) -> str:
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower().strip()


def _safe_decode(data: bytes) -> str:
    # utf-8 como default; com fallback tolerante
    try:
        return data.decode("utf-8")
    except Exception:
        return data.decode("utf-8", errors="replace")


def _extract_pdf(data: bytes) -> str:
    reader = PdfReader(BytesIO(data))
    parts: List[str] = []
    for i, page in enumerate(reader.pages):
        t = page.extract_text() or ""
        if t.strip():
            parts.append(f"[PDF:page={i+1}]\n{t.strip()}")
    return "\n\n".join(parts).strip()


def _extract_docx(data: bytes) -> str:
    doc = Document(BytesIO(data))
    parts = [p.text.strip() for p in doc.paragraphs if (p.text or "").strip()]
    return "\n".join(parts).strip()


def _extract_xlsx(data: bytes, max_sheets: int = 10, max_rows: int = 500, max_cols: int = 40) -> str:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet_names = wb.sheetnames[:max_sheets]

    out: List[str] = []
    for name in sheet_names:
        ws = wb[name]
        out.append(f"[XLSX:sheet={name}]")

        rcount = 0
        for row in ws.iter_rows(values_only=True):
            rcount += 1
            if rcount > max_rows:
                out.append("[... linhas truncadas ...]")
                break

            # limita colunas
            cells = row[:max_cols]
            # normaliza
            line = "\t".join("" if c is None else str(c) for c in cells).strip()
            if line:
                out.append(line)

        out.append("")  # separador

    return "\n".join(out).strip()


def extract_text_from_bytes(filename: str, data: bytes) -> str:
    ext = _get_ext(filename)
    if ext not in SUPPORTED_EXT:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Formato não suportado: .{ext or '?'} (suportados: {sorted(SUPPORTED_EXT)})",
        )

    if ext in {"txt", "md", "csv"}:
        return _safe_decode(data).strip()

    if ext == "pdf":
        return _extract_pdf(data)

    if ext == "docx":
        return _extract_docx(data)

    if ext == "xlsx":
        return _extract_xlsx(data)

    # fallback (não deve chegar aqui)
    return _safe_decode(data).strip()


async def extract_texts_from_uploads(
    files: List[UploadFile] | None,
    *,
    max_file_bytes: int = 8 * 1024 * 1024,   # 8 MB por arquivo (PoC)
    max_total_bytes: int = 20 * 1024 * 1024, # 20 MB total (PoC)
    max_chars_per_doc: int = 120_000,        # alinhado ao seu enforce_max_chars
) -> Tuple[List[str], List[str]]:
    """
    Retorna: (docs_text, warnings)
    - docs_text: lista de textos prontos para synthesize_matrix
    - warnings: avisos de truncamento / arquivos ignorados
    """
    docs_text: List[str] = []
    warnings: List[str] = []

    total = 0
    for f in files or []:
        data = await f.read()
        size = len(data)
        total += size

        if total > max_total_bytes:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="Total de arquivos excede limite do PoC. Envie menos arquivos ou divida em lotes.",
            )

        if size > max_file_bytes:
            warnings.append(f"{f.filename}: truncado para {max_file_bytes} bytes (PoC)")
            data = data[:max_file_bytes]

        text = extract_text_from_bytes(f.filename or "", data)

        if len(text) > max_chars_per_doc:
            warnings.append(f"{f.filename}: truncado para {max_chars_per_doc} chars (PoC)")
            text = text[:max_chars_per_doc]

        if text.strip():
            docs_text.append(f"[FILE:{f.filename}]\n{text.strip()}")

    return docs_text, warnings
